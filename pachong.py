#coding=utf-8

#抓取精品课网站中的课程，把有优惠券的课程筛选出来
#第一步：访问ke.youdao.com 获取精品课网页的所有的标签内容,例如：四六级，考研，实用英语...：
#第二步：访问标签页，获取课程详情页的url
#第三步：获取课程详情页需要的信息
#第四步：保存到Excel表中

import requests
import urllib3
import re
import sys
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook

#抓取标签,"http://ke.youdao.com"
def get_labels(url,label_file):
    urllib3.disable_warnings()
    resq = requests.get(url,verify=False).text
    labels = re.findall(r"href=\"(/tag/\d+)",resq)
    valid_labellink=[]
    for label in labels:
        valid_labellink.append(url+label)
    with open(label_file,"w") as fp:
        for i in set(valid_labellink):
            fp.writelines(i+"\n")
    print "labes url get done"

#通过标签抓取课程详情页url
def get_kelink(labellink_file,kelink_file):
    valid_kelink=[]
    with open(labellink_file) as fp:
        for line in fp:
            requests.packages.urllib3.disable_warnings()
            resq1=requests.get(line.strip(),verify=False).text
            ke_urls = re.findall(r"href=\"(https://ke\.youdao\.com/course/detail/\d+)",resq1)
            for kelink in ke_urls:
                valid_kelink.append(kelink.strip())
    with open(kelink_file,"w")as fp1:
        for kelink in set(valid_kelink):
            fp1.writelines(kelink+"\n")
    print "ke url get done"

#爬取课程名称，价格，开课时间，主讲老师
def get_courseinfo(kelink_file):
    result=[]
    with open(kelink_file) as fp:
            for keurl in fp:
                urllib3.disable_warnings()
                resq2 = requests.get(keurl.strip(),verify=False).text
                soup = BeautifulSoup(resq2,'html.parser')
                try:
                    #names = soup.select("div.info.info-without-video > h1")
                    names = soup.select("div.g-w.body > div > h1")
                    teachernames = soup.select("div.g-w.body > div > p")
                    coursetimes = soup.select("div.g-w.body > div > p")
                    if names!=None and teachernames!=None and coursetimes!=None:
                        data = {
                            'name': str(names[0]).strip().strip('<h1>').strip('</h1>'),
                            'teachername:': str(teachernames[0]).strip().strip("<p>").strip("</p>"),
                            'coursetimes': str(coursetimes[1]).strip().strip("<p>").strip("</p>"),
                            'url': keurl.strip()
                        }
                        result.append(data)
                        #print data
                    else:
                        print u"有属性为空了,skip skip"
                except Exception, e:
                        print e
            return result


#将爬虫下来的内容保存在Excel
def write_excel(filename,result):
    wb = load_workbook(filename)
    wb.guess_types = True
    ws=wb.active
    #excel表中有多少行，Excel的行和列是从第一行列开始的
    for i in range(1,len(result)+1):
        #取result列表中的每个data，每个data为一行
        result_item = result[i-1]
        #默认从第一列开始
        column_num=1
        #遍历字典data，每个属性增加一列
        for key,value1 in result_item.items():
            if value1.strip() != None:
                ws.cell(row=i,column=column_num,value=value1)
                column_num+=1
    wb.save(filename)


def main():
    '''
    #抓去ke.youdao.com上面的标签
    url="https://ke.youdao.com"
    label_filename="d:\\label.txt"
    get_labels(url,label_filename)
    #抓取课程详情页的url
    kelink_file="d:\\kelink.txt"
    get_kelink(label_filename,kelink_file)
    '''
    kelink_file = "d:\\kelink.txt"
    #抓取课程详细信息
    get_courseinfo(kelink_file)
    result = get_courseinfo(kelink_file)[1:10]
    write_excel("d:\\test.xlsx",result)

if __name__=="__main__":
    reload(sys)
    sys.setdefaultencoding("utf-8")
    main()



